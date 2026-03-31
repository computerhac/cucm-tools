import io
import csv

from fastapi import APIRouter, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, PatternFill

import axl
import database as db

router = APIRouter(prefix="/api/route-plan-audit", tags=["route-plan-audit"])

_CHUNK = 500  # max DIDs per SQL IN clause


# ---------------------------------------------------------------------------
# File parsing
# ---------------------------------------------------------------------------

def _parse_upload(content: bytes, filename: str, skip_header: bool) -> list[str]:
    numbers = []
    if filename.lower().endswith(".csv"):
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        for i, row in enumerate(reader):
            if skip_header and i == 0:
                continue
            if row and str(row[0]).strip():
                numbers.append(str(row[0]).strip())
    else:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        for i, row in enumerate(ws.iter_rows(min_col=1, max_col=1, values_only=True)):
            if skip_header and i == 0:
                continue
            val = row[0]
            if val is not None and str(val).strip():
                numbers.append(str(val).strip())
        wb.close()
    return numbers


# ---------------------------------------------------------------------------
# Sorting and range grouping
# ---------------------------------------------------------------------------

def _sort_key(num: str):
    digits = num.replace("\\", "").replace("+", "").replace("-", "").replace(" ", "")
    try:
        return (0, int(digits), num)
    except ValueError:
        return (1, 0, num)


def _group_ranges(numbers: list[str]) -> list[dict]:
    """Sort and mark where a new range group begins (gap > 1 or non-numeric)."""
    sorted_nums = sorted(dict.fromkeys(numbers), key=_sort_key)
    result = []
    prev_int = None
    for num in sorted_nums:
        digits = num.replace("\\", "").replace("+", "").replace("-", "").replace(" ", "")
        try:
            curr_int = int(digits)
            is_numeric = True
        except ValueError:
            curr_int = None
            is_numeric = False

        if not result:
            range_break = False
        elif prev_int is None or not is_numeric or curr_int - prev_int > 1:
            range_break = True
        else:
            range_break = False

        result.append({"number": num, "range_break": range_break})
        prev_int = curr_int if is_numeric else None
    return result


# ---------------------------------------------------------------------------
# CUCM lookup
# ---------------------------------------------------------------------------

def _lookup_cucm(host, port, username, password, verify_ssl, numbers: list[str]) -> dict[str, dict]:
    found: dict[str, dict] = {}

    for i in range(0, len(numbers), _CHUNK):
        chunk = numbers[i : i + _CHUNK]
        in_list = ", ".join(f"'{n.replace(chr(39), chr(39)*2)}'" for n in chunk)

        sql = f"""
            SELECT np.dnorpattern AS pattern,
                   np.description AS description,
                   CAST(np.tkpatternusage AS VARCHAR(20)) AS usage_type
            FROM numplan np
            WHERE np.dnorpattern IN ({in_list})
            ORDER BY np.dnorpattern
        """
        rows = axl.raw_query(host, port, username, password, verify_ssl, sql)
        for row in rows:
            pat = row.get("pattern", "").strip()
            if not pat:
                continue
            usage_code = row.get("usage_type", "").strip()
            if pat in found:
                found[pat]["count"] += 1
            else:
                found[pat] = {
                    "type": axl.PATTERN_USAGE.get(usage_code, f"Unknown ({usage_code})"),
                    "description": row.get("description", "").strip(),
                    "count": 1,
                }

        sql_vm = f"""
            SELECT vmp.directorynumber AS pattern,
                   vmp.description AS description
            FROM voicemessagingpilot vmp
            WHERE vmp.directorynumber IN ({in_list})
        """
        rows_vm = axl.raw_query(host, port, username, password, verify_ssl, sql_vm)
        for row in rows_vm:
            pat = row.get("pattern", "").strip()
            if pat and pat not in found:
                found[pat] = {
                    "type": "Voice Mail Pilot",
                    "description": row.get("description", "").strip(),
                    "count": 1,
                }

    for entry in found.values():
        if entry["count"] > 1:
            entry["type"] += f" ({entry['count']} partitions)"

    return found


# ---------------------------------------------------------------------------
# Endpoint
# ---------------------------------------------------------------------------

@router.post("/process")
async def process_audit(
    cluster_id: int = Form(...),
    skip_header: bool = Form(True),
    file: UploadFile = File(...),
):
    creds = db.get_cluster_credentials(cluster_id)
    if not creds:
        raise HTTPException(404, "Cluster not found.")

    filename = file.filename or ""
    if not (filename.lower().endswith(".xlsx") or filename.lower().endswith(".csv")):
        raise HTTPException(400, "Only .xlsx and .csv files are supported.")

    content = await file.read()
    try:
        numbers = _parse_upload(content, filename, skip_header)
    except Exception as e:
        raise HTTPException(400, f"Failed to parse file: {e}")

    if not numbers:
        raise HTTPException(400, "No numbers found in column A.")

    try:
        cucm_data = _lookup_cucm(
            host=creds["host"], port=creds["port"],
            username=creds["username"], password=creds["password"],
            verify_ssl=creds["verify_ssl"],
            numbers=numbers,
        )
    except Exception as e:
        raise HTTPException(500, f"CUCM query failed: {e}")

    grouped = _group_ranges(numbers)
    found_count = sum(1 for item in grouped if item["number"] in cucm_data)

    # Build output workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Route Plan Audit"

    ws.append(["Number", "Type", "Description"])
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
    ws.freeze_panes = "A2"

    grey = PatternFill("solid", fgColor="F3F4F6")

    for item in grouped:
        if item["range_break"]:
            ws.append(["", "", ""])

        num = item["number"]
        match = cucm_data.get(num)
        ws.append([num, match["type"] if match else "", match["description"] if match else ""])

        if not match:
            idx = ws.max_row
            for col in range(1, 4):
                ws.cell(row=idx, column=col).fill = grey

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 52

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    total = len(grouped)
    not_found = total - found_count

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=route_plan_audit.xlsx",
            "X-Audit-Total":     str(total),
            "X-Audit-Found":     str(found_count),
            "X-Audit-Not-Found": str(not_found),
            "Access-Control-Expose-Headers": "X-Audit-Total, X-Audit-Found, X-Audit-Not-Found",
        },
    )
